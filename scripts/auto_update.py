"""Trickcal Dashboard auto-update — lightweight version for cron.

Does:
  1. Extract revenue CSV from xlsx (fast, ~10s)
  2. Check inspection data folder for NEW inspection points (incremental)
  3. Git commit & push if changes exist

Does NOT:
  - Re-extract ALL inspection xlsx every time (too slow)
  - Access Google Sheets (requires user login via CDP)
"""
import openpyxl, csv, json, os, sys, pathlib, subprocess, datetime

DASHBOARD_DIR = pathlib.Path("C:/Users/user/Desktop/trickcal_dashboard")
DATA_DIR = DASHBOARD_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
XLSX_PATH = pathlib.Path("C:/Users/user/Desktop/전달문서/통합 문서1.xlsx")
INSPECTION_BASE = pathlib.Path("C:/Users/user/Desktop/트릭컬 종합 문서/8_점검데이터")

def extract_revenue():
    if not XLSX_PATH.exists():
        print(f"[WARN] Revenue xlsx not found: {XLSX_PATH}")
        return False
    
    wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True, data_only=True)
    
    # Global
    ws = wb["26년 매출"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c) if c else "" for c in rows[0]]
    col_map = {h: i for i, h in enumerate(header)}
    
    global_data = []
    for row in rows[1:]:
        if not row[0]: continue
        ds = row[col_map.get("ds", 0)]
        if hasattr(ds, "strftime"): ds = ds.strftime("%Y-%m-%d")
        global_data.append({
            "date": str(ds),
            "active_accounts": int(row[col_map.get("活跃账号数", 1)] or 0),
            "paid_active": int(row[col_map.get("活跃的付费账号数", 2)] or 0),
            "new_accounts": int(row[col_map.get("新增账号数", 3)] or 0),
            "revenue_usd": int(row[col_map.get("总收入_美元", 4)] or 0),
            "paying_accounts": int(row[col_map.get("付费账号数", 5)] or 0),
            "pay_rate": float(row[col_map.get("活跃账号付费率", 6)] or 0),
            "arpu": float(row[col_map.get("活跃ARPU", 7)] or 0),
            "arppu": float(row[col_map.get("活跃ARPPU", 8)] or 0),
        })
    
    out = DATA_DIR / "global_revenue.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=global_data[0].keys())
        w.writeheader(); w.writerows(global_data)
    print(f"[OK] Global: {len(global_data)} rows")
    
    # JP
    try:
        ws2 = wb["JP"]
        rows2 = list(ws2.iter_rows(values_only=True))
        header2 = [str(c) if c else "" for c in rows2[0]]
        col_map2 = {h: i for i, h in enumerate(header2)}
        
        jp_data = []
        for row in rows2[1:]:
            if not row[0] or not row[col_map2.get("ds")]: continue
            ds = row[col_map2["ds"]]
            if hasattr(ds, "strftime"): ds = ds.strftime("%Y-%m-%d")
            try:
                si = lambda v: int(v or 0) if isinstance(v, (int, float)) else 0
                sf = lambda v: float(v or 0) if isinstance(v, (int, float)) else 0.0
                jp_data.append({
                    "date": str(ds),
                    "platform": str(row[col_map2.get("b_platform", 1)] or ""),
                    "country": str(row[col_map2.get("country", 2)] or ""),
                    "active_accounts": si(row[col_map2.get("活跃账号数", 3)]),
                    "paid_active": si(row[col_map2.get("活跃的付费账号数", 4)]),
                    "new_accounts": si(row[col_map2.get("新增账号数", 5)]),
                    "active_characters": si(row[col_map2.get("活跃角色数", 6)]),
                    "revenue_usd": si(row[col_map2.get("总收入_美元", 7)]),
                    "paying_accounts": si(row[col_map2.get("付费账号数", 8)]),
                    "pay_rate": sf(row[col_map2.get("活跃账号付费率", 9)]),
                    "arpu": sf(row[col_map2.get("活跃ARPU", 10)]),
                })
            except Exception as e:
                print(f"[WARN] JP skip: {e}")
        
        out2 = DATA_DIR / "jp_revenue.csv"
        with open(out2, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=jp_data[0].keys())
            w.writeheader(); w.writerows(jp_data)
        print(f"[OK] JP: {len(jp_data)} rows")
    except Exception as e:
        print(f"[WARN] JP sheet error: {e}")
    
    wb.close()
    return True

def check_new_inspections():
    """Check if there are new inspection folders not yet in inspection_metrics.json"""
    metrics_file = DATA_DIR / "inspection_metrics.json"
    existing_dates = set()
    if metrics_file.exists():
        with open(metrics_file, "r", encoding="utf-8") as f:
            existing = json.load(f)
        existing_dates = {d["date"] for d in existing}
    
    if not INSPECTION_BASE.exists():
        print("[INFO] Inspection base not found")
        return False
    
    subdirs = sorted([d for d in INSPECTION_BASE.iterdir() if d.is_dir()])
    new_folders = []
    
    for subdir in subdirs:
        folder_name = subdir.name
        date_str = ""
        for ch in folder_name.replace(" ", ""):
            if ch.isdigit():
                date_str += ch
            else:
                break
        if len(date_str) >= 4:
            formatted_date = f"2026-{date_str[:2]}-{date_str[2:4]}"
            if formatted_date not in existing_dates:
                xlsx_count = len(list(subdir.glob("**/*.xlsx")))
                if xlsx_count > 0:
                    new_folders.append((formatted_date, subdir.name, xlsx_count))
    
    if new_folders:
        print(f"[INFO] New inspection data found: {[(d[0], d[2]) for d in new_folders]}")
        print("[INFO] Run full inspection extraction separately (too slow for cron)")
        return True
    return False

def git_push():
    os.chdir(str(DASHBOARD_DIR))
    subprocess.run(["git", "add", "-A"], check=True, capture_output=True)
    result = subprocess.run(["git", "diff", "--cached", "--quiet"], capture_output=True)
    if result.returncode == 0:
        print("[INFO] No changes to commit")
        return False
    today = datetime.date.today().strftime("%Y-%m-%d")
    msg = f"data: auto update ({today})"
    subprocess.run(["git", "commit", "-m", msg], check=True, capture_output=True)
    subprocess.run(["git", "push", "origin", "main"], check=True, capture_output=True)
    print(f"[OK] Git push: {msg}")
    return True

if __name__ == "__main__":
    print(f"=== Dashboard Auto-Update {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} ===")
    changed = extract_revenue()
    changed = check_new_inspections() or changed
    git_push()
    print("=== Done ===")
